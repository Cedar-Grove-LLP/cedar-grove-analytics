#!/usr/bin/env python3
"""Verify the Invoices (2026) workbook's month-tab matrix against the RAW
attorney timesheets — the last layer of the verification loop.

The workbook computes each matrix cell as
    QUERY(timesheet month tab A10:E, SUM(hours) WHERE client) x client rate
    (+ SUM(adjustment col D) for Sam McClure)
and each attorney's Billable Earnings (rate table) is IMPORTRANGE(timesheet B3).

This script recomputes both from the timesheet exports and diffs against the
extracted snapshot (scripts/invoices-real-snapshot.json). READ-ONLY everywhere.

Usage:
  python3 scripts/verify-timesheets.py <exports-dir>
where <exports-dir> holds "2026 - Invoices (<Surname>).xlsx" files.
Export the timesheets and the 2026 workbook at the same time — the timesheets
are live, so a stale snapshot will show drift, not logic errors.
"""

import json
import os
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SNAPSHOT = os.path.join(ROOT, "scripts", "invoices-real-snapshot.json")

# Matrix short name -> timesheet file surname + full-name fragment for the rate table.
ATTORNEYS = {
    "Sam": ("McClure", "mcclure"),
    "Colin": ("van Loon", "van loon"),
    "Michael O": ("Ohta", "ohta"),
    "Michael": ("Ohta", "ohta"),
    "Molly": ("Manning", "manning"),
    "Michael L": ("Levin", "levin"),
    "Valery": ("Uscanga", "uscanga"),
    "David": ("Popkin", "popkin"),
    "Nick": ("Agate", "agate"),
    "Paige": ("Wilson", "wilson"),
    "Martyna": ("Skrodzka", "skrodzka"),
}
MONTH_TAB = {m: m.capitalize() for m in
             ["january", "february", "march", "april", "may", "june", "july"]}
EPS = 0.02


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def load_timesheets(folder):
    books = {}
    for short, (surname, _) in ATTORNEYS.items():
        path = os.path.join(folder, f"2026 - Invoices ({surname}).xlsx")
        if surname not in books and os.path.exists(path):
            books[surname] = openpyxl.load_workbook(path, data_only=True)
    return books


def read_timesheet_tab(ws):
    """Parse one timesheet month tab, handling both layouts.

    Legacy (Jan-era):  entry cols A=Client B=Date C=Hours D=Billables Earnings;
                       workbook imports earnings as B2 (+B5 83(b) fee).
    Current (Jun-era): entry cols A=Client B=Date C=Hours D=Adjustment E=Earnings;
                       workbook imports earnings as B3.
    Detection: the entry-header row is the row whose col A is 'Client'; the
    layout is 'current' iff its col D label contains 'Adjust'.
    """
    header_row = None
    for r in range(1, 16):
        if str(ws.cell(row=r, column=1).value).strip() == "Client":
            header_row = r
            break
    if header_row is None:
        return None
    col_d_label = str(ws.cell(row=header_row, column=4).value or "").lower()
    has_adjustment = "adjust" in col_d_label
    has_billable_to_client = "billable to client" in col_d_label

    # Summary block (col A labels / col B values, rows 1..8)
    summary = {}
    for r in range(1, 9):
        label = str(ws.cell(row=r, column=1).value or "").strip()
        if label:
            summary[label] = num(ws.cell(row=r, column=2).value)
    earnings = summary.get("Billable Earnings", 0.0)
    fee83b = next((v for k, v in summary.items() if k.startswith("83(b)")), 0.0)
    imported_earnings = earnings if has_adjustment else earnings + fee83b

    hours, adjustments, billable_to_client = defaultdict(float), defaultdict(float), defaultdict(float)
    for r in range(header_row + 1, ws.max_row + 1):
        client = ws.cell(row=r, column=1).value
        if client is None:
            continue
        key = str(client).strip()
        hours[key] += num(ws.cell(row=r, column=3).value)
        if has_adjustment:
            adjustments[key] += num(ws.cell(row=r, column=4).value)
        if has_billable_to_client:
            billable_to_client[key] += num(ws.cell(row=r, column=4).value)
    return {
        "hours": hours, "adjustments": adjustments,
        "billableToClient": billable_to_client if has_billable_to_client else None,
        "importedEarnings": imported_earnings,
        "earnings": earnings, "fee83b": fee83b,
    }


def main():
    folder = sys.argv[1] if len(sys.argv) > 1 else None
    if not folder or not os.path.isdir(folder):
        sys.exit(__doc__)
    with open(SNAPSHOT) as f:
        wb = json.load(f)
    books = load_timesheets(folder)
    print(f"snapshot: {wb['source']} ({wb['extractedOn']}) · timesheets: {sorted(books)}\n")

    total_cells = total_diff = total_b3 = b3_diff = 0
    for mkey, month in wb["months"].items():
        matrix = month.get("matrix")
        if not matrix:
            continue
        table = month.get("attorneyTable", {})
        theaders = table.get("headers", [])
        be_idx = theaders.index("Billable Earnings") if "Billable Earnings" in theaders else -1
        month_report = []

        for j, short in enumerate(matrix["attorneys"]):
            surname, frag = ATTORNEYS[short]
            book = books.get(surname)
            if not book or MONTH_TAB[mkey] not in book.sheetnames:
                continue
            ts = read_timesheet_tab(book[MONTH_TAB[mkey]])
            if ts is None:
                continue

            # client rate from the month's attorney rate table (first value column)
            trow = next((r for r in table.get("rows", []) if frag in r["name"].lower()), None)
            if not trow:
                continue
            rate = num(trow["vals"][0])

            # per-client matrix cells vs timesheet recomputation. Where the
            # timesheet has a 'Billable to Client' column (Valery), the workbook
            # switched formulas between months — accept either hours×rate or
            # the explicit billable amounts.
            for row in matrix["rows"]:
                client = row["client"].strip()
                candidates = [ts["hours"].get(client, 0.0) * rate]
                if short == "Sam":
                    candidates[0] += ts["adjustments"].get(client, 0.0)
                if ts["billableToClient"] is not None:
                    candidates.append(ts["billableToClient"].get(client, 0.0))
                actual = row["billings"][j]
                if abs(actual) < EPS and all(abs(c) < EPS for c in candidates):
                    continue
                total_cells += 1
                if all(abs(actual - c) > EPS for c in candidates):
                    total_diff += 1
                    month_report.append(f"    DIFF {short:10s} {client[:34]:34s} matrix={actual:>10.2f} timesheet={candidates[0]:>10.2f}")

            # timesheet Billable Earnings vs the workbook's rate-table import.
            # Layouts differ per attorney: the import may be B3-only or B2+B5
            # (earnings + 83(b)) — accept either; flag when NEITHER matches.
            if be_idx >= 0:
                imported = trow["vals"][be_idx]
                if isinstance(imported, (int, float)):
                    total_b3 += 1
                    cands = {round(ts["earnings"], 2), round(ts["earnings"] + ts["fee83b"], 2)}
                    if all(abs(c - imported) > EPS for c in cands):
                        b3_diff += 1
                        month_report.append(f"    EARN-DIFF {short:10s} timesheet={ts['earnings']:.2f} (+83b {ts['fee83b']:.2f}) vs workbook import={imported:.2f}")
                # surface the workbook's own Check/Diff flag when nonzero
                if "Diff" in theaders:
                    d = trow["vals"][theaders.index("Diff")]
                    if isinstance(d, (int, float)) and abs(d) > EPS:
                        month_report.append(f"    SHEET-FLAG {short:10s} workbook's own Diff column = {d:.2f} (Check formula disagrees with import)")

        status = "all match" if not month_report else f"{len(month_report)} issue(s)"
        print(f"{MONTH_TAB[mkey]:10s} {status}")
        for line in month_report[:12]:
            print(line)
        if len(month_report) > 12:
            print(f"    ... and {len(month_report) - 12} more")

    print(f"\nmatrix cells compared: {total_cells}, diffs: {total_diff}")
    print(f"B3 earnings imports compared: {total_b3}, diffs: {b3_diff}")
    print("PASS" if total_diff == 0 and b3_diff == 0 else "DIFFS FOUND (drift or workbook bug — see above)")


if __name__ == "__main__":
    main()
