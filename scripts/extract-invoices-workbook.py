#!/usr/bin/env python3
"""Extract the real Jan-June dataset for the dashboard's "Invoices (testing)" tab
from an .xlsx export of the Cedar Grove "Invoices (2026)" Google Sheet.

READ-ONLY on the workbook. Writes two files:
  scripts/invoices-real-snapshot.json   — canonical snapshot (used for drift diffs)
  src/utils/invoicesRealData.mjs        — generated module the tab + tests import

Usage:
  python3 scripts/extract-invoices-workbook.py "/path/to/Cedar Grove LLP - Invoices (2026).xlsx"

Re-run with a fresh export to refresh the "Real (Jan-Jun)" mode; the script
prints a drift report of every value that changed since the last snapshot.
"""

import json
import os
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Both outputs are gitignored — they hold REAL attorney-compensation figures and
# stay local-only. The committed src/utils/invoicesRealData.mjs is a SYNTHETIC
# stand-in; regenerate it from this local .mjs with
#   node scripts/make-synthetic-invoices-data.mjs
SNAPSHOT = os.path.join(ROOT, "scripts", "invoices-real-snapshot.json")
MJS_OUT = os.path.join(ROOT, "src", "utils", "invoicesRealData.local.mjs")

MONTHS = ["january", "february", "march", "april", "may", "june", "july"]
SHEET_NAME = {m: m.capitalize() for m in MONTHS}
# Frozen snapshot tabs, captured separately so the calendar-month index mapping stays clean.
EXTRA_MONTHS = {"june-original": "June - original"}
CASH_MONTHS = ["january", "february", "march", "april", "may", "june"]  # cash rows 3..8

# Month-tab matrix: header row 20 tail columns → dataset field names.
MATRIX_FIELD = {
    "Sum Billables": "sumBillables", "83(b) Elections": "elections83b", "Filing Fees": "filingFees",
    "Fees Notes": "feesNotes", "Outside Counsel": "outsideCounsel", "Outside Counsel Notes": "ocNotes",
    "Prior Deferred": "priorDeferred", "Prior Deferral Toggle": "priorToggle",
    "Deferred This Month": "deferredThisMonth", "Total Deferred": "totalDeferred",
    "Write Off": "writeOff", "Invoiced": "invoiced", "General Notes": "generalNotes",
    "Contact Name": "contactName", "Contact Email": "contactEmail", "Payment Terms": "paymentTerms",
}
MATRIX_NUMERIC = {"sumBillables", "elections83b", "filingFees", "outsideCounsel", "priorDeferred",
                  "deferredThisMonth", "totalDeferred", "writeOff", "invoiced"}
MATRIX_TEXT = {"feesNotes", "ocNotes", "priorToggle", "generalNotes", "contactName", "contactEmail"}

# Monthly tab column B rows 2..16, in order (matches WATERFALL_ROWS keys).
WF_KEYS = [
    "gross", "writeOffs", "attorneyBillables", "flatFee83b", "filingFees",
    "outsideCounsel", "netAccrued", "deferred", "revenueAccrued", "cgfDonation",
    "revenueMinusCgf", "attorneyPayout", "netRevenueBeforeOpEx", "opEx", "firmProfits",
]
# Which of those are INPUTS to computeMonthlyWaterfall (the rest are derived).
WF_INPUT_KEYS = [
    "attorneyBillables", "flatFee83b", "filingFees", "outsideCounsel",
    "writeOffs", "deferred", "attorneyPayout", "opEx",
]

# P&L line rows (1-indexed sheet rows) → dataset keys. Values cols B..G (Jan-Jun).
PNL_LINE_ROWS = {
    "revenue": 4,
    "software": 9,
    "malpractice": 12,
    "franchiseTaxes": 13,
    "filingFees": 14,
    "reimbursements": 17,
    "misc": 20,
    "outsideCounsel": 23,
    "attorneys": 26,
    "payrollTaxes": 38,
    "charitable": 43,
    "cedarGrove": 44,
}
PNL_CONSULTANT_ROWS = {
    "Valyria": 29, "Valery Uscanga": 30, "Martyna Skrodzka": 31, "Nick Agate": 32,
    "David Popkin": 33, "Paige Wilson": 34, "Accountants": 35,
}
PNL_SHEET_ROWS = {"totalRevenue": 5, "totalExpenses": 40, "cgfTotal": 45, "netIncome": 47}

# Note-cell fill colors used on the Profits Paid (Sam) ledger.
FILL_HIGHLIGHT = {"D9EAD3": "green", "FFF2CC": "tan"}


def num(v):
    return round(float(v), 6) if isinstance(v, (int, float)) else 0.0


def cell_highlight(cell):
    try:
        if cell.fill and cell.fill.fill_type == "solid":
            rgb = str(cell.fill.fgColor.rgb or "")
            return FILL_HIGHLIGHT.get(rgb[-6:].upper(), "")
    except Exception:
        pass
    return ""


def display_note(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):  # datetime note cells display as e.g. "March 2026"
        return v.strftime("%B %Y")
    return str(v)


def raw_val(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(float(v), 6)
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v) if v is not None else None


def extract_month_tab(ws):
    """Waterfall + attorney rate table + client billings matrix of a month tab."""
    col_b, errors = {}, {}
    for r, k in enumerate(WF_KEYS, start=2):
        v = ws.cell(row=r, column=2).value
        if isinstance(v, str) and v.startswith("#"):
            errors[k] = v  # the SHEET cell itself is an error (e.g. broken IMPORTRANGE)
            col_b[k] = 0.0
        else:
            col_b[k] = num(v)
    month = {"inputs": {k: col_b[k] for k in WF_INPUT_KEYS}, "sheet": col_b, "sheetErrors": errors}

    # Attorney rate table — header row 1, anchored at the 'Attorney' cell.
    start = None
    for c in range(2, 30):
        if ws.cell(row=1, column=c).value == "Attorney":
            start = c
            break
    table = {"headers": [], "rows": []}
    if start:
        c = start + 1
        while ws.cell(row=1, column=c).value is not None:
            table["headers"].append(str(ws.cell(row=1, column=c).value))
            c += 1
        r = 2
        while ws.cell(row=r, column=start).value is not None:
            table["rows"].append({
                "name": str(ws.cell(row=r, column=start).value),
                "vals": [raw_val(ws.cell(row=r, column=start + 1 + k).value) for k in range(len(table["headers"]))],
            })
            r += 1
    month["attorneyTable"] = table

    # Client billings matrix — header row 20, data rows 21+. Rows where every
    # numeric column is zero are dropped (they add nothing to any total).
    headers = []
    c = 1
    while ws.cell(row=20, column=c).value is not None:
        headers.append(str(ws.cell(row=20, column=c).value))
        c += 1
    rows, total_rows = [], 0
    if "Sum Billables" in headers:
        sb = headers.index("Sum Billables")
        attorneys = headers[1:sb]
        r = 21
        while r <= ws.max_row and ws.cell(row=r, column=1).value is not None:
            total_rows += 1
            billings = [num(ws.cell(row=r, column=2 + j).value) for j in range(len(attorneys))]
            obj = {"client": str(ws.cell(row=r, column=1).value), "billings": billings}
            for k in range(sb, len(headers)):
                field = MATRIX_FIELD.get(headers[k])
                if not field:
                    continue
                v = ws.cell(row=r, column=k + 1).value
                if field in MATRIX_NUMERIC:
                    obj[field] = num(v)
                elif field in MATRIX_TEXT:
                    obj[field] = display_note(v) if v is not None else ""
                else:
                    obj[field] = raw_val(v)
            if any(billings) or any(obj.get(f) for f in MATRIX_NUMERIC):
                rows.append(obj)
            r += 1
        month["matrix"] = {"attorneys": attorneys, "rows": rows, "totalRows": total_rows}
    return month


def extract_register(ws):
    register = []
    for r in range(2, ws.max_row + 1):
        client = ws.cell(row=r, column=1).value
        amount = ws.cell(row=r, column=2).value
        if client is None and amount is None:
            continue
        y = ws.cell(row=r, column=3).value
        register.append({
            "client": str(client or ""),
            "amount": num(amount),
            "year": int(y) if isinstance(y, (int, float)) else None,
            "dateSent": raw_val(ws.cell(row=r, column=4).value),
            "status": str(ws.cell(row=r, column=5).value or ""),
            "lastReminder": raw_val(ws.cell(row=r, column=6).value),
            "dateReceived": raw_val(ws.cell(row=r, column=7).value),
            "notes": str(ws.cell(row=r, column=8).value or ""),
        })
    return register


def extract(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {
        "source": os.path.basename(xlsx_path),
        "extractedOn": date.today().isoformat(),
        "months": {},
        "monthsExtra": {},
        "cash": {},
        "pnl": {"lines": {}, "consultants": {}, "sheet": {}},
        "paymentTotal": 0.0,
    }

    for m in MONTHS:
        out["months"][m] = extract_month_tab(wb[SHEET_NAME[m]])
    for key, sheet_name in EXTRA_MONTHS.items():
        out["monthsExtra"][key] = extract_month_tab(wb[sheet_name])

    ws = wb["Cash Accounting Summary"]
    for i, m in enumerate(CASH_MONTHS):
        r = 3 + i
        q = ws.cell(row=r, column=8).value
        out["cash"][m] = {
            "inputs": {
                "cashReceived": num(ws.cell(row=r, column=2).value),
                "expenses": num(ws.cell(row=r, column=3).value),
                "cgfDonation": num(ws.cell(row=r, column=4).value),
                "attorneyPayout": num(ws.cell(row=r, column=5).value),
            },
            "sheet": {
                "profits": num(ws.cell(row=r, column=6).value),
                "revenue": num(ws.cell(row=r, column=7).value),
                "qRevenue": num(q) if isinstance(q, (int, float)) else None,
            },
        }

    ws = wb["P&L"]
    def prow(r):
        return [num(ws.cell(row=r, column=2 + c).value) for c in range(6)]
    for key, r in PNL_LINE_ROWS.items():
        out["pnl"]["lines"][key] = prow(r)
    for key, r in PNL_CONSULTANT_ROWS.items():
        out["pnl"]["consultants"][key] = prow(r)
    for key, r in PNL_SHEET_ROWS.items():
        out["pnl"]["sheet"][key] = prow(r)

    out["paymentTotal"] = num(wb["Payment Status"].cell(row=1, column=2).value)

    # Profits Paid (Sam) — manual ledger, headers on row 2, data from row 3.
    ws = wb["Profits Paid (Sam)"]
    ledger = []
    r = 3
    while True:
        d = ws.cell(row=r, column=1).value
        if d is None:
            break
        note_cell = ws.cell(row=r, column=4)
        ledger.append({
            "date": d.strftime("%m-%d-%Y") if hasattr(d, "strftime") else str(d),
            "description": str(ws.cell(row=r, column=2).value or ""),
            "amount": num(ws.cell(row=r, column=3).value),
            "note": display_note(note_cell.value),
            "highlight": cell_highlight(note_cell),
        })
        r += 1
    out["profitsPaid"] = ledger

    def iso(v):
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else (str(v) if v else None)

    def raw(v):
        if isinstance(v, (int, float)):
            return round(float(v), 6)
        return str(v) if v is not None else None

    # Rate Sheet — levels A1..P2, rows 2..21 (Cravath notes column excluded).
    ws = wb["Rate Sheet"]
    rate_rows = []
    for r in range(2, 22):
        level = ws.cell(row=r, column=1).value
        if level is None:
            break
        g = ws.cell(row=r, column=7).value
        rate_rows.append({
            "level": str(level),
            "tier": str(ws.cell(row=r, column=2).value or ""),
            "clientRate": num(ws.cell(row=r, column=3).value),
            "attorneyRate": num(ws.cell(row=r, column=4).value),
            "colinRate": raw(ws.cell(row=r, column=5).value),
            "salary": raw(ws.cell(row=r, column=6).value),
            "cravath": num(g) if isinstance(g, (int, float)) else None,
        })
    out["rateSheet"] = rate_rows

    # Expenses V2 — vendor rows: category, label, Jan–Dec, P&L tag (col O).
    ws = wb["Expenses V2"]
    exp_rows = []
    for r in range(2, 200):
        cat = ws.cell(row=r, column=1).value
        if cat is None:
            continue
        exp_rows.append({
            "category": str(cat),
            "label": str(ws.cell(row=r, column=2).value or ""),
            "vals": [num(ws.cell(row=r, column=3 + m).value) for m in range(12)],
            "pnlCat": (str(ws.cell(row=r, column=15).value) if ws.cell(row=r, column=15).value is not None else None),
        })
    out["expenses"] = exp_rows

    # Balance Sheet — label / current period / note rows (1..40).
    ws = wb["Balance Sheet"]
    bal_rows = []
    for r in range(1, 41):
        label = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        note = ws.cell(row=r, column=3).value
        if label is None and val is None and note is None:
            continue
        bal_rows.append({"row": r, "label": str(label) if label is not None else "", "value": raw(val), "note": raw(note)})
    out["balanceSheet"] = bal_rows

    # Payment Status — the full invoice register (no payment-terms column;
    # terms live in the clients sheet / Firestore) + the frozen 06/30 copy.
    out["paymentRegister"] = extract_register(wb["Payment Status"])
    out["paymentRegisterCopy"] = extract_register(wb["0630 Copy of Payment Status"])
    out["paymentTotalCopy"] = num(wb["0630 Copy of Payment Status"].cell(row=1, column=2).value)
    return out


def flatten(d, prefix=""):
    flat = {}
    for k, v in d.items():
        p = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            flat.update(flatten(v, p))
        elif isinstance(v, list):
            for i, x in enumerate(v):
                flat[f"{p}[{i}]"] = x
        else:
            flat[p] = v
    return flat


def drift_report(old, new):
    skip = {"source", "extractedOn"}
    fo, fn = flatten(old), flatten(new)
    changes = [
        (k, fo.get(k), fn.get(k))
        for k in sorted(set(fo) | set(fn))
        if k.split(".")[0] not in skip and fo.get(k) != fn.get(k)
    ]
    if not changes:
        print("No drift — snapshot matches the previous extraction.")
        return
    print(f"DRIFT: {len(changes)} value(s) changed since the last snapshot:")
    for k, o, n in changes:
        print(f"  {k}: {o} -> {n}")


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    xlsx = sys.argv[1]
    if not os.path.exists(xlsx):
        sys.exit(f"Not found: {xlsx}")

    new = extract(xlsx)

    if os.path.exists(SNAPSHOT):
        with open(SNAPSHOT) as f:
            drift_report(json.load(f), new)
    else:
        print("No previous snapshot — writing the first one.")

    with open(SNAPSHOT, "w") as f:
        json.dump(new, f, indent=1)
    print(f"Wrote {os.path.relpath(SNAPSHOT, ROOT)}")

    header = (
        "// GENERATED by scripts/extract-invoices-workbook.py — do not edit by hand.\n"
        f"// Source: {new['source']} (extracted {new['extractedOn']}).\n"
        "// Real Jan–June figures captured read-only from the Invoices (2026) workbook;\n"
        "// consumed by the self-contained \"Invoices (testing)\" dashboard tab and its tests.\n"
    )
    with open(MJS_OUT, "w") as f:
        f.write(header + "export const REAL_WORKBOOK = " + json.dumps(new, indent=1) + ";\n")
    print(f"Wrote {os.path.relpath(MJS_OUT, ROOT)}")


if __name__ == "__main__":
    main()
